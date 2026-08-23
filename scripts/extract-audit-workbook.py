#!/usr/bin/env python3
"""
extract-audit-workbook.py - one-time-ish extractor.

Turns the team's manual workbook into a machine-readable input file. Two jobs:

  1. Supply the domain list and each site's declared email configuration, which
     is what the DNS checker needs as input.
  2. Carry the team's RECORDED Pass/Fail cells forward, so an automated check
     can be compared against a year of human judgement instead of silently
     replacing it. That comparison is the point; see compare mode in
     fleet-email-dns.py.

Requires openpyxl (read-only, one-time). The DNS checker itself does not.

Usage:
  ./scripts/extract-audit-workbook.py --xlsx path/to/audit.xlsx --out data/fleet-email-inventory.json
"""
import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("needs openpyxl:  pip install openpyxl")

# HEADER TEXT, not column position. Row 1 is a merged group header, row 2 the
# real header, data starts row 3.
#
# THIS USED TO BE A DICT OF INDICES AND IT WENT WRONG. The workbook is a live
# file people edit. Between the 2026-08-18 import and 2026-08-23 it gained three
# columns -- GCDN Configured, WAF Mode, Admin Routes Protected by IP Access Rule
# -- inserted BEFORE Notes. `notes` was pinned to index 27; index 27 is now
# `GCDN Configured` and Notes moved to 30. Re-running this script would have
# imported GCDN values into the notes field on all 78 sites and exited 0.
#
# Matching on header text cannot drift when a column is inserted, and a RENAMED
# or deleted column now fails loudly instead of silently reading its neighbour.
# Same rule as everywhere else here: a confident-looking value standing in for
# the wrong thing is the bug this project keeps making.
HEADERS = {
    "domain": "WordPress Domain",
    "hosted": "Hosted",
    "dns": "DNS",
    "smtp_plugin": "WP Plugin",
    "provider": "Provider",
    "sending_domain": "Email Sending Domain",
    "envelope_from": "Envelope From Address",
    "from_address": "From Address",
    "env_from_match": "Envelope/From  Match",
    "from_name": "From Name",
    "spf": "SPF",
    "dkim": "DKIM",
    "dmarc": "DMARC",
    "tracking": "Tracking",
    "receiving": "Receiving",
    "single_cm_user": "Single  CM User",
    "php_version": "PHP Version",
    "wp_version": "WordPress Version",
    "themes_utd": "Themes Up to Date",
    "plugins_utd": "Plugins  Up to Date w/ Inactive Removed",
    "activity_log": "WP Activity  Log Plugin  Configured",
    "wp_2fa": "WP 2FA  Plugin  Configured",
    "hide_login": "WP Hide  Login Plugin Configured",
    "xmlrpc_disabled": "xmlrpc.php Disabled",
    "php_blocker": "Dev/Test Sites Use PHP Blocker (Pantheon only)",
    "keeper_password": "Generated Password in Keeper",
    "wp2shell_remedied": "wp2shell Security Flaw Remedied?",
    "notes": "Notes",
}

# Columns present in the sheet on 2026-08-23 that this script does not read.
# Listed so a future reader can tell "we decided not to import this" from
# "nobody noticed it existed".
IGNORED_HEADERS = (
    "GCDN Configured (Pantheon only)",
    "WAF Mode",
    "Admin Routes Protected by IP Access Rule",
)

# Filled by resolve_columns() from the header row. Never hardcoded again.
C = {}


def _norm(v):
    """Header cells carry double spaces and stray newlines. Compare on shape."""
    return " ".join(str(v or "").split()).lower()


def resolve_columns(header_row):
    """Map field name -> column index by matching header TEXT.

    Hard error on anything missing. A partial map would produce a file that
    looks like a successful import and is quietly wrong in one column, which is
    exactly what the positional version did.
    """
    found = {}
    want = {_norm(v): k for k, v in HEADERS.items()}
    for i, cellv in enumerate(header_row):
        k = want.get(_norm(cellv))
        if k is not None and k not in found:
            found[k] = i
    missing = sorted(set(HEADERS) - set(found))
    if missing:
        seen = [str(c) for c in header_row if c]
        sys.exit("workbook header does not match. Missing: %s\nHeader row was: %s"
                 % (", ".join(missing), " | ".join(seen)))
    unread = [str(c) for c in header_row
              if c and _norm(c) not in want
              and _norm(c) not in {_norm(x) for x in IGNORED_HEADERS}]
    if unread:
        print("note: %d column(s) in the sheet are not imported and are not on "
              "the known-ignored list: %s" % (len(unread), ", ".join(unread)),
              file=sys.stderr)
    return found

# Recorded verdicts we will compare against. Blank stays blank; it is NOT a
# pass and it is NOT a fail. Same rule as the ledger: unknown is a value.
VERDICT_FIELDS = ("spf", "dkim", "dmarc", "tracking", "receiving", "env_from_match")


def cell(row, key):
    i = C[key]
    if i >= len(row):
        return None
    v = row[i]
    if v is None:
        return None
    s = str(v).replace("\n", " ").strip()
    return s if s else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    rows = list(wb["Sites"].iter_rows(values_only=True))
    # Row 2 (index 1) is the real header; row 1 is a merged group label.
    C.update(resolve_columns(rows[1]))
    sites = []
    for r in rows[2:]:
        d = cell(r, "domain")
        if not d:
            continue
        rec = {"domain": d.lower(), "hosted": cell(r, "hosted"), "dns": cell(r, "dns")}
        for k in ("smtp_plugin", "provider", "sending_domain", "envelope_from",
                  "from_address", "from_name", "notes"):
            rec[k] = cell(r, k)
        rec["recorded"] = {k: cell(r, k) for k in VERDICT_FIELDS}
        # Security columns come along for the ride; they are the seed of the
        # inventory file and cost nothing to carry.
        rec["recorded_security"] = {
            k: cell(r, k) for k in (
                "single_cm_user", "php_version", "wp_version", "themes_utd",
                "plugins_utd", "activity_log", "wp_2fa", "hide_login",
                "xmlrpc_disabled", "keeper_password", "wp2shell_remedied")
        }
        sites.append(rec)

    plugins = []
    if "Security Plugins" in wb.sheetnames:
        for r in list(wb["Security Plugins"].iter_rows(values_only=True))[1:]:
            if r and r[0]:
                plugins.append({"name": str(r[0]).strip(),
                                "url": str(r[1]).strip() if r[1] else None,
                                "role": str(r[2]).strip() if r[2] else None,
                                "notes": str(r[3]).strip() if r[3] else None})

    out = {"source": a.xlsx.split("/")[-1], "site_count": len(sites),
           "sites": sites, "security_plugins": plugins}
    with open(a.out, "w") as fh:
        json.dump(out, fh, indent=1, sort_keys=True)
    print("extracted %d sites, %d security plugins -> %s" % (len(sites), len(plugins), a.out))


if __name__ == "__main__":
    main()
